import openpyxl
import cv2
from openpyxl.styles import PatternFill
from colormap.colors import rgb2hex
from openpyxl.utils import get_column_letter
import numpy as np


def rgb_to_hex(rgb_color_list):
    rgb_color_list = np.ndarray.tolist(rgb_color_list)
    for i in range(len(rgb_color_list[0])):
        for j in range(len(rgb_color_list)):
            r = rgb_color_list[j][i][2]
            g = rgb_color_list[j][i][1]
            b = rgb_color_list[j][i][0]

            rgb_color_list[j][i] = rgb2hex(r, g, b)[1:7]
    return rgb_color_list


def image_to_excel_transform(excel_file_name, color_list, width_scale=1):
    # Creating excel file
    book = openpyxl.Workbook()
    # Select sheet
    sheet = book.active
    # Bypassing cells of sheet
    for columns in range(1, len(color_list[0])):
        for rows in range(1, len(color_list)):
            # Changing cells width
            sheet.column_dimensions[get_column_letter(columns)].width = 2.5 * width_scale  # Column width of Excel sheet
            # Coloring cells by coordinates
            sheet.cell(row=rows, column=columns).fill = PatternFill("solid", start_color=color_list[rows][columns])
    # excel file saving in select directory
    book.save('{}.xlsx'.format(excel_file_name))
    book.close()


def main():
    image_name = 'python.jpg'
    image = cv2.imread(image_name)
    image_hex_list = rgb_to_hex(image)
    image_to_excel_transform(image_name, image_hex_list)


if __name__ == '__main__':
    main()
